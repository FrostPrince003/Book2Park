from tkinter import *
from tkinter import messagebox
from openpyxl import *
import cv2
import pickle
import cvzone
import numpy as np
import os



root=Tk()
root.title("ParkiN")
root.geometry('925x500+300+200')
root.configure(bg='#fff')
root.resizable(False,False)

def book():
    vehicleno = user.get()
    phoneno = Phone.get()
    gmailac = gmail.get()
    
    if vehicleno =='' and  phoneno =='' and gmailac == '':
        messagebox.showerror("X",'Please fill the details')
    
       
    building=Toplevel(root)
    building.title('Parkin!')
    building.geometry('925x500+300+200')
    building.configure(bg='#fff')
    building.resizable(False,False)   
    Label(building,text='Please select Your Parking Space',fg='#57a1f8',bg='white',font =('Microsoft Yahei UI Light',20)).place(x=285,y=5)
            
    Parking_A = Button(building,text='Parking A',pady=1,width=10,fg='black',border=0,bg='#FFFFEB',font=('Microsoft Yahei UI Light',14,),command = footage)
    Parking_A.place(x=140,y=220)

    
    
    park1 = PhotoImage(file='parkinga.png')
    park11=Label(building,image=park1,border=0,bg='white')
    park11.grid(row=4,column=10,rowspan=2,columnspan=2,padx=40,pady=10,sticky=W)
    park11.image = park1
    
    

    
    Parking_B = Button(building,text='Parking B',pady=1,width=10,fg='black',border=0,bg='#FFFFEB',font=('Microsoft Yahei UI Light',14,),command = build2)
    Parking_B.place(x=650,y=220)
    
    park2 = PhotoImage(file='parkingb.png')
    park22=Label(building,image=park2,border=0,bg='white')
    park22.grid(row=4,column=50,rowspan=5,columnspan=5,padx=40,pady=10,sticky=W)
    park22.image = park2
    
    
    Parking_C = Button(building,text='Parking C',pady=1,width=10,fg='black',border=0,bg='#FFFFEB',font=('Microsoft Yahei UI Light',14,),command = build3)
    Parking_C.place(x=140,y=420)
    
    Parking_D = Button(building,text='Parking D',pady=1,width=10,fg='black',border=0,bg='#FFFFEB',font=('Microsoft Yahei UI Light',14,),command = build4)
    Parking_D.place(x=650,y=420)

    
    

img = PhotoImage(file='parking1.png')
Label(root,image=img,border=0,bg='white').place(x=-350,y=-265)

frame=Frame(root,width=400,height=390,bg='#fff')
frame.place(x=450,y=80)

heading=Label(frame,text='Welcome to Parkin!',fg="#57a1f8",bg="white",font=('Microsoft Yahei UI Light',20,))
heading.place(x=120, y=5)

def insert():

	# if user not fill any entry
	# then print "empty input"
	if (Phone.get() == "" or
		user.get() == "" or
		gmail.get() == "" ):

		print("Please fill all details")

	else:

		# assigning the max row and max column
		# value upto which data is written
		# in an excel sheet to the variable
		current_row = sheet.max_row
		current_column = sheet.max_column

		# get method returns current text
		# as string which we write into
		# excel spreadsheet at particular location
		sheet.cell(row=current_row + 1, column=1).value = Phone.get()
		sheet.cell(row=current_row + 1, column=2).value = user.get()
		sheet.cell(row=current_row + 1, column=3).value = gmail.get()


		# save the file
		wb.save('Book1.xlsx')

		# set focus on the name_field box
		user.focus_set()

		# call the clear() function
  
def footage():
    cap = cv2.VideoCapture('carPark.mp4')
    
    file_path = os.path.join('D:', 'Coding', 'C course', '.vscode', 'CarParkingPos')
    



    with open(file_path, 'rb') as f:
        posList = pickle.load(f)

    width,height = 107,46

    def checkParkingSpace(imgPro):
        spaceCounter=0

        for pos in posList:
            x,y = pos

            imgCrop = imgPro[y:y+height,x:x+width]
    
            #cv2.imshow(str(x*y),imgCrop)
            count = cv2.countNonZero(imgCrop)#The number in front of the boxes
            if count < 900 :
                 color = (0,255,0)
                 thickness = 5
            else:
                 color = (0,0,255)
                 thickness = 2
    
            cv2.rectangle(img, pos, (pos[0] + width, pos[1] + height), color, 2)
            cvzone.putTextRect(img, str(count), (x, y + height - 3), scale=0.8, thickness=1, offset=0)
    
    
            cvzone.putTextRect(img, str(count), (x, y + height - 3), scale=0.8, thickness=1, offset=0)
    
    while True:
        
        if cap.get(cv2.CAP_PROP_POS_FRAMES) == cap.get(cv2.CAP_PROP_FRAME_COUNT):
            cap.set(cv2.CAP_PROP_POS_FRAMES,0)
        success, img = cap.read()
        imgGray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
        imgBlur = cv2.GaussianBlur(imgGray,(3,3),1)
        imgThreshold = cv2.adaptiveThreshold(imgBlur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,cv2.THRESH_BINARY_INV,25,16)
        imgMedian = cv2.medianBlur(imgThreshold,5)
        kernel = np.ones((3, 3), np.uint8)
        imgDilate = cv2.dilate(imgMedian,kernel,iterations=1)
    
        checkParkingSpace(imgDilate)
        cv2.imshow("Image",img)
        cv2.imshow("ImageBlur",imgBlur)
        cv2.imshow("ImageThres",imgMedian)
        cv2.waitKey(50)
     
    
    
    
   
    
    
		

ok=Button(frame,text='Book',pady=1,width=25,fg='white',border=0,bg='#57a1f8',font=('Microsoft Yahei UI Light',14),command = lambda:[insert(),book()] )
ok.place(x=115,y=250)

def on_enter(e):
    user.delete(0,'end')
def on_leave(e):
    if user.get()=='':
        user.insert(0,'Vehicle Number ')


user=Entry(frame,width=25,fg="black",border=0,bg='white',font=('Microsoft Yahei UI Light',13))
user.place(x=110,y=67)

# This input will be given by the parking coordinator

user.insert(0,'Vehicle Number')
user.bind("<FocusIn>",on_enter)
user.bind("<FocusOut>",on_leave)

Frame(frame,width=295,height=2,bg='black').place(x=110,y=90)

def on_enter(e):
    Phone.delete(0,'end')
def on_leave(e):
    if Phone.get()=='':
        Phone.insert(0,'Phone No.')

Phone=Entry(frame,width=25,fg="black",border=0,bg='white',font=('Microsoft Yahei UI Light',13))
Phone.place(x=110,y=134)
Phone.insert(0,'Phone No.')
Phone.bind("<FocusIn>",on_enter)
Phone.bind("<FocusOut>",on_leave)

Frame(frame,width=295,height=2,bg='black').place(x=110,y=160)

def on_enter(e):
    gmail.delete(0,'end')
def on_leave(e):
    if gmail.get()=='':
        gmail.insert(0,'Enter your Gmail')

gmail=Entry(frame,width=25,fg="black",border=0,bg='white',font=('Microsoft Yahei UI Light',13))
gmail.place(x=110,y=200)
gmail.insert(0,'Enter your Gmail ')
gmail.bind("<FocusIn>",on_enter)
gmail.bind("<FocusOut>",on_leave)

Frame(frame,width=295,height=2,bg='black').place(x=110,y=230)

# For Parking A
#def build1():
#    park_a=Toplevel(root)
#    park_a.title('Parkin!')
#    park_a.geometry('925x500+300+200')
#    park_a.configure(bg='#fff')
#    park_a.resizable(False,False)
    
    
    
# For Parking B
def build2():
    park_b=Toplevel(root)
    park_b.title('Parkin!')
    park_b.geometry('925x500+300+200')
    park_b.configure(bg='#fff')
    park_b.resizable(False,False)
    
# For Parking C
def build3():
    park_c=Toplevel(root)
    park_c.title('Parkin!')
    park_c.geometry('925x500+300+200')
    park_c.configure(bg='#fff')
    park_c.resizable(False,False)

# For Parking D
def build4():
    park_d=Toplevel(root)
    park_d.title('Parkin!')
    park_d.geometry('925x500+300+200')
    park_d.configure(bg='#fff')
    park_d.resizable(False,False)

# create the sheet object

wb = load_workbook('Book1.xlsx')
sheet = wb.active

def excel():

	# resize the width of columns in
	# excel spreadsheet
	sheet.column_dimensions['A'].width = 20
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 30


	# write given data to an excel spreadsheet
	# at particular location
	sheet.cell(row=1, column=1).value = "mobile no."
	sheet.cell(row=1, column=2).value = "Vehicle No."
	sheet.cell(row=1, column=3).value = "e-mail id"




root.mainloop()










































