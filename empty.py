from tkinter import *
from tkinter import messagebox
from openpyxl import *
from main import *
from datetime import datetime

root = Tk()
root.title("ParkiN")
root.geometry('925x500+300+200')
root.configure(bg='#fff')
root.resizable(False, False)


def book():

    vehicleno = user.get()
    phoneno = Phone.get()
    gmailac = gmail.get()

    if vehicleno == '' and phoneno == '' and gmailac == '':
        messagebox.showerror("X", 'Please fill the details')

    building = Toplevel(root)
    building.title('Parkin!')
    building.geometry('925x500+300+200')
    building.configure(bg='#fff')
    building.resizable(False, False)
    Label(building, text='Please select Your Parking Space', fg='#57a1f8', bg='white',
          font=('Microsoft Yahei UI Light', 20)).place(x=285, y=5)

    Parking_A = Button(building, text='Parking A', pady=1, width=10, fg='black', border=0, bg='#FFFFEB',
                       font=('Microsoft Yahei UI Light', 14,), command=footage)
    Parking_A.place(x=195, y=290)

    park1 = PhotoImage(file='parkinga.png')
    park11 = Label(building, image=park1, border=0, bg='white')
    park11.place(x=150 , y=80)
    park11.image = park1

    Parking_B = Button(building, text='Parking B', pady=1, width=10, fg='black', border=0, bg='#FFFFEB',
                       font=('Microsoft Yahei UI Light', 14,), command=build2)
    Parking_B.place(x=620, y=290)

    park2 = PhotoImage(file='parkingb.png')
    park22 = Label(building, image=park2, border=0, bg='white')
    park22.place(x=595,y=80)
    park22.image = park2

    Parking_C = Button(building, text='Parking C', pady=1, width=10, fg='black', border=0, bg='#FFFFEB',
                       font=('Microsoft Yahei UI Light', 14,), command=build3)
    Parking_C.place(x=140, y=420)

    Parking_D = Button(building, text='Parking D', pady=1, width=10, fg='black', border=0, bg='#FFFFEB',
                       font=('Microsoft Yahei UI Light', 14,), command=build4)
    Parking_D.place(x=650, y=420)






img = PhotoImage(file='parking1.png')
Label(root, image=img, border=0, bg='white').place(x=-350, y=-265)

frame = Frame(root, width=400, height=390, bg='#fff')
frame.place(x=450, y=80)

heading = Label(frame, text='Welcome to Parkin!', fg="#57a1f8", bg="white", font=('Microsoft Yahei UI Light', 20,))
heading.place(x=120, y=5)


def insert():
    # if user not fill any entry
    # then print "empty input"
    if (Phone.get() == "Phone No." or
            user.get() == "Vehicle Number" or
            gmail.get() == "Enter your Gmail"):
        messagebox.showerror("Enter all details","Please fill all details")

    else:

        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = Phone.get()
        sheet.cell(row=current_row + 1, column=2).value = user.get()
        sheet.cell(row=current_row + 1, column=3).value = gmail.get()
        sheet.cell(row=current_row + 1, column=4).value = datetime.now()

        # save the file
        wb.save('book1.xlsx')

        # set focus on the name_field box
        user.focus_set()
        book()

    # call the clear() function




ok = Button(frame, text='Book', pady=1, width=25, fg='white', border=0, bg='#57a1f8',
            font=('Microsoft Yahei UI Light', 14), command=lambda: [insert()])
ok.place(x=115, y=250)


def on_enter(e):
    user.delete(0, 'end')


def on_leave(e):
    if user.get() == '':
        user.insert(0, 'Vehicle Number ')


user = Entry(frame, width=25, fg="black", border=0, bg='white', font=('Microsoft Yahei UI Light', 13))
user.place(x=110, y=67)

# This input will be given by the parking coordinator

user.insert(0, 'Vehicle Number')
user.bind("<FocusIn>", on_enter)
user.bind("<FocusOut>", on_leave)

Frame(frame, width=295, height=2, bg='black').place(x=110, y=90)


def on_enter(e):
    Phone.delete(0, 'end')


def on_leave(e):
    if Phone.get() == '':
        Phone.insert(0, 'Phone No.')


Phone = Entry(frame, width=25, fg="black", border=0, bg='white', font=('Microsoft Yahei UI Light', 13))
Phone.place(x=110, y=134)
Phone.insert(0, 'Phone No.')
Phone.bind("<FocusIn>", on_enter)
Phone.bind("<FocusOut>", on_leave)

Frame(frame, width=295, height=2, bg='black').place(x=110, y=160)


def on_enter(e):
    gmail.delete(0, 'end')


def on_leave(e):
    if gmail.get() == '':
        gmail.insert(0, 'Enter your Gmail')


gmail = Entry(frame, width=25, fg="black", border=0, bg='white', font=('Microsoft Yahei UI Light', 13))
gmail.place(x=110, y=200)
gmail.insert(0, 'Enter your Gmail ')
gmail.bind("<FocusIn>", on_enter)
gmail.bind("<FocusOut>", on_leave)

Frame(frame, width=295, height=2, bg='black').place(x=110, y=230)


# For Parking A
# def build1():
#    park_a=Toplevel(root)
#    park_a.title('Parkin!')
#    park_a.geometry('925x500+300+200')
#    park_a.configure(bg='#fff')
#    park_a.resizable(False,False)


# For Parking B
def build2():
    park_b = Toplevel(root)
    park_b.title('Parkin!')
    park_b.geometry('925x500+300+200')
    park_b.configure(bg='#fff')
    park_b.resizable(False, False)


# For Parking C
def build3():
    park_c = Toplevel(root)
    park_c.title('Parkin!')
    park_c.geometry('925x500+300+200')
    park_c.configure(bg='#fff')
    park_c.resizable(False, False)


# For Parking D
def build4():
    park_d = Toplevel(root)
    park_d.title('Parkin!')
    park_d.geometry('925x500+300+200')
    park_d.configure(bg='#fff')
    park_d.resizable(False, False)


# create the sheet object

wb = load_workbook('book1.xlsx')
sheet = wb.active


def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 30

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "mobile no."
    sheet.cell(row=1, column=2).value = "Vehicle No."
    sheet.cell(row=1, column=3).value = "e-mail id"

root.mainloop()