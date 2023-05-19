
# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *

# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook('Book1.xlsx')

# create the sheet object
sheet = wb.active


def excel():
	
	# resize the width of columns in
	# excel spreadsheet
	sheet.column_dimensions['A'].width = 20
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 30


	# write given data to an excel spreadsheet
	# at particular location
	sheet.cell(row=1, column=1).value = "mobile no."
	sheet.cell(row=1, column=2).value = "Vehicle No."
	sheet.cell(row=1, column=3).value = "e-mail id"

# Function to set focus (cursor)
def focus1(event):
	# set focus on the course_field box
	name_field.focus_set()


# Function to set focus
def focus2(event):
	# set focus on the sem_field box
	vn_field.focus_set()


# Function to set focus
def focus3(event):
	# set focus on the form_no_field box
	time_field.focus_set()

def focus4(event):
	# set focus on the form_no_field box
	submit.focus_set()

# Function for clearing the
# contents of text entry boxes
def clear():
	
	# clear the content of text entry box
	name_field.delete(0, END)
	vn_field.delete(0, END)
	time_field.delete(0, END)
	


# Function to take data from GUI
# window and write to an excel file
def insert():
	
	# if user not fill any entry
	# then print "empty input"
	if (name_field.get() == "" or
		vn_field.get() == "" or
		time_field.get() == "" ):
			
		print("Please fill all details")

	else:

		# assigning the max row and max column
		# value upto which data is written
		# in an excel sheet to the variable
		current_row = sheet.max_row
		current_column = sheet.max_column

		# get method returns current text
		# as string which we write into
		# excel spreadsheet at particular location
		sheet.cell(row=current_row + 1, column=1).value = name_field.get()
		sheet.cell(row=current_row + 1, column=2).value = vn_field.get()
		sheet.cell(row=current_row + 1, column=3).value = time_field.get()
		

		# save the file
		wb.save('Book1.xlsx')

		# set focus on the name_field box
		name_field.focus_set()

		# call the clear() function
		clear()





	
# create a GUI window
root = Tk()

# set the background colour of GUI window
root.configure(background='light green')

# set the title of GUI window
# root.title("registration form")

# set the configuration of GUI window
root.geometry("925x500")
#image frame
frame=Frame(root,width=400,height=400,bg='#fff')
frame.place(x=450,y=20) 
img = PhotoImage(file='parking1.png')
Label(frame,image=img,border=0,bg='white').place(x=-390,y=-340)


    

excel()

# create a Form label
heading = Label(root, text="Welcome to ParkIN", font=('Microsoft Yahei UI Light',18) , bg="light green")
sub_heading = Label(root, text="Book To ParkIn", font=('Microsoft Yahei UI Light',16) , bg="light green")

# create a Name label
name = Label(root, text="Mobile No.:", bg="light green", font=('Microsoft Yahei UI Light',14) )

	# create a Course label
vn = Label(root, text="Vehicle no.", bg="light green", font=('Microsoft Yahei UI Light',14) )

time = Label(root, text="Email :", bg="light green", font=('Microsoft Yahei UI Light',14) )

	# grid method is used for placing
	# the widgets at respective positions
	# in table like structure .
heading.grid(row=1, column=1)
sub_heading.grid(row=2, column=1)
name.grid(row=3, column=0)
vn.grid(row=4, column=0)
time.grid(row=5, column=0)
	

	# create a text entry box
	# for typing the information
name_field = Entry(root)
vn_field = Entry(root)
time_field = Entry(root)

	# bind method of widget is used for
	# the binding the function with the events

	# whenever the enter key is pressed
	# then call the focus1 function
name_field.bind("<Return>", focus2)

	# whenever the enter key is pressed
	# then call the focus2 function
vn_field.bind("<Return>", focus3)

	# whenever the enter key is pressed
	# then call the focus3 function
time_field.bind("<Return>", focus4)



	
	# grid method is used for placing
	# the widgets at respective positions
	# in table like structure .
name_field.grid(row=3, column=1, ipadx="100")
vn_field.grid(row=4, column=1, ipadx="100")
time_field.grid(row=5, column=1, ipadx="100")

	# call excel function
excel()

	# create a Submit Button and place into the root window
submit = Button(root, text="Submit", fg="Black", bg="Red", command=insert, font=("Microsoft Yahei UI Light",12))
submit.grid(row=8, column=1)


	# start the GUI
root.mainloop()


